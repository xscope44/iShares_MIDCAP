from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from alive_progress import alive_bar
import config as c

c.prCyan(f"******** Add Analysis sheet: {c.ishares_out_prefix} ********")
ishares_out_file = c.find_newest_file_simple(c.data_folder_path, c.ishares_out_prefix, c.file_ishares_out_extension)

# Load the existing file
existing_workbook = load_workbook(ishares_out_file)
existing_sheet = existing_workbook[c.ishares_gurufocus_sheet_name]
ishares_sheet = existing_workbook[c.ishares_out_sheet_name]
# Check if the new sheet name already exists
if c.analysis_sheet_name in existing_workbook.sheetnames:
    print(f"The sheet '{c.analysis_sheet_name}' already exists in the file.")
    # You can handle the situation here, such as renaming or deleting the existing sheet
else:
# Create a new sheet
    new_sheet = existing_workbook.create_sheet(title=c.analysis_sheet_name)

    # Copy the first three columns from the first sheet to the new sheet
    for row in ishares_sheet.iter_rows(max_row=ishares_sheet.max_row, max_col=3, values_only=True):
        new_sheet.append(row)

    #### Add Google Price: ##########################

    # Get the column letter for the new column
    new_column_letter = get_column_letter(new_sheet.max_column + 1)

    # Set the header for the new column
    new_sheet[new_column_letter + "1"] = c.price_column_name

    # Apply the formula to the new column
    
    for row in range(2, 3):
        # formula = f'=INDEX(GoogleFinance(A{row},"price",TODAY()),2,2)'
        formula = f'=IFNA(INDEX(GoogleFinance(A{row}; "close"; TODAY()); 2; 2); IFNA(INDEX(GoogleFinance(A{row}; "close"; TODAY()-1); 2; 2); IFNA(INDEX(GoogleFinance(A{row}; "close"; TODAY()-2); 2; 2); IFNA(INDEX(GoogleFinance(A{row}; "close"; TODAY()-3); 2; 2); IFNA(INDEX(GoogleFinance(A{row}; "close"; TODAY()-4); 2; 2); INDEX(GoogleFinance(A{row}; "close"; TODAY()-5); 2; 2))))))'
        cell = new_sheet[new_column_letter + str(row)]
        cell.value = formula

    #### Add 6M Price: ##########################
    # Get the column letter for the new column
    new_column_letter = get_column_letter(new_sheet.max_column + 1)

    # Set the header for the new column
    new_sheet[new_column_letter + "1"] = c.six_month_price_column_name

    # Apply the formula to the new column
    for row in range(2, 3):
        formula = f'=IFNA(INDEX(GoogleFinance(A{row};"close";TODAY()-180);2;2); IFNA(INDEX(GoogleFinance(A{row};"close";TODAY()-181);2;2); IFNA(INDEX(GoogleFinance(A{row};"close";TODAY()-179);2;2); INDEX(GoogleFinance(A{row};"close";TODAY()-182);2;2))))'
        
        cell = new_sheet[new_column_letter + str(row)]
        cell.value = formula

    #### Add 6M RPS: ##########################
    # Get the column letter for the new column
    new_column_letter = get_column_letter(new_sheet.max_column + 1)

    # Set the header for the new column
    new_sheet[new_column_letter + "1"] = c.six_month_rps_column_name

    # Apply the formula to the new column
    for row in range(2, new_sheet.max_row + 1):
        formula = f'=H{row}/E{row}-1'
        cell = new_sheet[new_column_letter + str(row)]
        cell.value = formula


    #### Add Upside Potencial: GF Value or Fair Value vs current price ##########################
    # Get the column letter for the new column
    new_column_letter = get_column_letter(new_sheet.max_column + 1)

    # Set the header for the new column
    new_sheet[new_column_letter + "1"] = c.upside_potencial_column_name

    # Apply the formula to the new column
    for row in range(2, new_sheet.max_row + 1):
        formula = f'=J{row}/H{row}-1'
        cell = new_sheet[new_column_letter + str(row)]
        cell.value = formula

    #### Add ETF Price collumn: ##########################
    # Get the column letter for the 9th column in the first sheet
    column_letter_9th = get_column_letter(9)

    # Copy the 9th column from the first sheet to the new sheet
    # Get the column letter for the new column
    new_column_letter = get_column_letter(new_sheet.max_column + 1)
    new_sheet[new_column_letter + "1"] = "ETF Price"
    for row in range(1, ishares_sheet.max_row + 1):
        cell = ishares_sheet[column_letter_9th + str(row)]
        if cell.value == 'Price':
            new_sheet.cell(row=row, column=new_sheet.max_column, value='ETF Price')
        else:    
            new_sheet.cell(row=row, column=new_sheet.max_column, value=cell.value)

    #### Add GuruFocus collums: ##########################
    existing_sheet = existing_workbook[c.ishares_gurufocus_sheet_name]
        # Get the column letter for the 9th column in the first sheet
    max_col = existing_sheet.max_column
    # print(max_col)
    with alive_bar(max_col-2, force_tty=True) as bar:
        for i in range(2,max_col):
            column_letter_2th = get_column_letter(i)
            # Copy the 9th column from the first sheet to the new sheet
            # Get the column letter for the new column
            new_column_letter = get_column_letter(new_sheet.max_column + 1)
            new_sheet[new_column_letter + "1"] = "ETF Price"
            for row in range(1, existing_sheet.max_row + 1):
                cell = existing_sheet[column_letter_2th + str(row)]
                new_sheet.cell(row=row, column=new_sheet.max_column, value=cell.value)
            bar()

    # Save the changes to the existing file
    existing_workbook.save(ishares_out_file)
    existing_workbook.close()
    