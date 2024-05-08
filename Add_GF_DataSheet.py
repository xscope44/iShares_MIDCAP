from openpyxl import load_workbook
import config as c

c.prCyan(f"******** Add Gurufocus sheet to: {c.ishares_out_prefix} ********")
file_gurufocus = c.find_newest_file_simple(c.data_folder_path, c.gurufocus_prefix, c.file_ishares_out_extension)
file_ishares_out = c.find_newest_file_simple(c.data_folder_path, c.ishares_out_prefix, c.file_ishares_out_extension)

# Load the source file
source_workbook = load_workbook(file_gurufocus, read_only=True)
source_sheet = source_workbook[c.gf_sheet_name]

# Load the existing file
existing_workbook = load_workbook(file_ishares_out)
existing_sheet_names = existing_workbook.sheetnames

# Check if the new sheet name already exists in the existing file
if c.ishares_gurufocus_sheet_name in existing_sheet_names:
    # Remove the existing sheet before renaming
    existing_workbook.remove(existing_workbook[c.ishares_gurufocus_sheet_name])

# Create a new sheet in the existing file
existing_workbook.create_sheet(c.ishares_gurufocus_sheet_name)
existing_sheet = existing_workbook[c.ishares_gurufocus_sheet_name]

# Copy the source sheet data to the existing sheet
for row in source_sheet.iter_rows(values_only=True):
    existing_sheet.append(row)

# Save the changes to the existing file
existing_workbook.save(file_ishares_out)
existing_workbook.close()
